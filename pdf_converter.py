from __future__ import annotations

import argparse
import csv
import ctypes
import json
import os
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

from PIL import Image, ImageOps, ImageSequence
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas


IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tif", ".tiff", ".webp"}
OFFICE_EXTS = {
    ".doc",
    ".docx",
    ".rtf",
    ".odt",
    ".ppt",
    ".pptx",
    ".odp",
    ".xls",
    ".xlsx",
    ".xlsm",
    ".ods",
    ".csv",
}
TEXT_EXTS = {".txt", ".md", ".log"}
SPREADSHEET_TABLE_EXTS = {".xlsx", ".xlsm", ".csv"}
SUPPORTED_EXTS = IMAGE_EXTS | OFFICE_EXTS | TEXT_EXTS | {".pdf"}
COM_CONVERSION_LOCK = threading.Lock()
SHUTDOWN_EVENT = threading.Event()
ACTIVE_PROCESS_LOCK = threading.Lock()
ACTIVE_SUBPROCESSES: set[subprocess.Popen] = set()
ACTIVE_EXTERNAL_PIDS: set[int] = set()
OFFICE_PROCESS_NAMES = {
    "WINWORD.EXE",
    "POWERPNT.EXE",
    "EXCEL.EXE",
    "WPS.EXE",
    "WPP.EXE",
    "ET.EXE",
    "SOFFICE.EXE",
    "SOFFICE.BIN",
}

TYPE_PROFILES = {
    "All supported files": SUPPORTED_EXTS,
    "Word documents": {".doc", ".docx", ".rtf", ".odt"},
    "PowerPoint files": {".ppt", ".pptx", ".odp"},
    "Excel and CSV files": {".xls", ".xlsx", ".xlsm", ".ods", ".csv"},
    "Images": IMAGE_EXTS,
    "Text and Markdown": TEXT_EXTS,
    "Existing PDFs": {".pdf"},
}

LANGUAGES = {"zh": "中文", "en": "English"}
PROFILE_LABELS = {
    "zh": {
        "All supported files": "全部支持格式",
        "Word documents": "Word 文档",
        "PowerPoint files": "PPT 演示文稿",
        "Excel and CSV files": "Excel / CSV 表格",
        "Images": "图片",
        "Text and Markdown": "文本 / Markdown",
        "Existing PDFs": "已有 PDF",
    },
    "en": {
        "All supported files": "All supported files",
        "Word documents": "Word documents",
        "PowerPoint files": "PowerPoint files",
        "Excel and CSV files": "Excel and CSV files",
        "Images": "Images",
        "Text and Markdown": "Text and Markdown",
        "Existing PDFs": "Existing PDFs",
    },
}

TEXT = {
    "zh": {
        "title": "PDF 转换器",
        "add_files": "添加文件",
        "add_folder": "添加文件夹",
        "clear": "清空",
        "include_subfolders": "包含子文件夹",
        "merge_final": "合并最终 PDF",
        "start_merged": "转换后合并",
        "convert_type": "转换类型",
        "merged_name": "合并后的 PDF 名称",
        "output_folder": "输出文件夹",
        "browse": "浏览",
        "move_up": "上移",
        "move_down": "下移",
        "remove": "删除",
        "split_pdf": "拆分 PDF",
        "start": "开始转换",
        "language": "语言",
        "drop_hint": "请把文件拖到列表区域，或使用上方按钮添加。",
        "selected": "已选择 {count} 个项目。",
        "choose_files": "选择要转换的文件",
        "choose_folder": "选择文件夹",
        "choose_output": "选择输出文件夹",
        "all_files": "所有文件",
        "no_files_title": "还没有文件",
        "no_files_body": "请先添加或拖入文件。",
        "starting": "开始转换...",
        "finished_title": "转换完成",
        "finished_body": "成功 {ok} 个，失败 {fail} 个。",
        "done_status": "完成：成功 {ok} 个，失败 {fail} 个。",
        "ok": "成功",
        "fail": "失败",
        "split_choose_pdf": "选择要拆分的 PDF",
        "split_range_label": "页码范围",
        "split_range_placeholder": "请输入页码范围，例如：1-3,5,8-10",
        "split_loading": "正在生成预览，请稍候...",
        "split_no_pages_selected": "请先选择页面，或输入页码范围。",
        "split_done": "已导出 {count} 页：{path}",
        "split_failed": "拆分失败：{error}",
        "split_no_pdf": "请先在列表中选中一个 PDF，或接下来选择一个 PDF 文件。",
        "split_preview_failed": "无法生成预览，将改用页码输入。",
        "export_selected": "导出选中页面",
        "select_all": "全选",
        "clear_selection": "清空选择",
        "back": "返回",
        "cancel": "取消",
        "page_label": "第 {page} 页",
        "add_failed_title": "添加失败",
        "add_failed_body": "有 {count} 个项目未添加。可能是不支持的格式，或不符合当前转换类型。",
        "add_failed_log": "未添加：{path}",
        "add_none": "没有可添加的文件。请检查格式或当前转换类型。",
    },
    "en": {
        "title": "PDF Converter",
        "add_files": "Add files",
        "add_folder": "Add folder",
        "clear": "Clear",
        "include_subfolders": "Include subfolders",
        "merge_final": "Merge final PDFs",
        "start_merged": "Convert and merge",
        "convert_type": "Convert type",
        "merged_name": "Merged PDF name",
        "output_folder": "Output folder",
        "browse": "Browse",
        "move_up": "Move up",
        "move_down": "Move down",
        "remove": "Remove",
        "split_pdf": "Split PDF",
        "start": "Start converting",
        "language": "Language",
        "drop_hint": "Drop files here or add them with the buttons.",
        "selected": "{count} item(s) selected.",
        "choose_files": "Choose files to convert",
        "choose_folder": "Choose a folder",
        "choose_output": "Choose output folder",
        "all_files": "All files",
        "no_files_title": "No files",
        "no_files_body": "Add or drop files first.",
        "starting": "Starting conversion...",
        "finished_title": "Finished",
        "finished_body": "{ok} succeeded, {fail} failed.",
        "done_status": "Done: {ok} succeeded, {fail} failed.",
        "ok": "OK",
        "fail": "FAIL",
        "split_choose_pdf": "Choose a PDF to split",
        "split_range_label": "Page range",
        "split_range_placeholder": "Enter pages, e.g. 1-3,5,8-10",
        "split_loading": "Creating previews, please wait...",
        "split_no_pages_selected": "Select pages or enter a page range first.",
        "split_done": "Exported {count} page(s): {path}",
        "split_failed": "Split failed: {error}",
        "split_no_pdf": "Select one PDF in the list, or choose a PDF next.",
        "split_preview_failed": "Preview could not be created, falling back to page range input.",
        "export_selected": "Export selected pages",
        "select_all": "Select all",
        "clear_selection": "Clear selection",
        "back": "Back",
        "cancel": "Cancel",
        "page_label": "Page {page}",
        "add_failed_title": "Add failed",
        "add_failed_body": "{count} item(s) were not added. They may be unsupported or excluded by the current convert type.",
        "add_failed_log": "Not added: {path}",
        "add_none": "No usable files were added. Check the format or current convert type.",
    },
}


@dataclass
class ConvertResult:
    source: Path
    ok: bool
    message: str
    output: Path | None = None


@dataclass
class StagedOutput:
    result: ConvertResult
    staged_pdf: Path
    final_target: Path


def safe_pdf_name(path: Path, output_dir: Path) -> Path:
    target = output_dir / f"{path.stem}.pdf"
    if not target.exists():
        return target

    i = 2
    while True:
        candidate = output_dir / f"{path.stem}_{i}.pdf"
        if not candidate.exists():
            return candidate
        i += 1


def enable_high_dpi() -> None:
    if os.name != "nt":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


def config_path() -> Path:
    base = os.environ.get("APPDATA")
    if base:
        return Path(base) / "PDFConverter" / "settings.json"
    return Path.home() / ".pdf_converter_settings.json"


def load_settings() -> dict[str, str]:
    path = config_path()
    if not path.exists():
        return {"language": "zh"}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"language": "zh"}
    language = data.get("language", "zh")
    if language not in LANGUAGES:
        language = "zh"
    return {"language": language}


def save_settings(settings: dict[str, str]) -> None:
    path = config_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")


def default_output_dir() -> Path:
    if os.name == "nt":
        try:
            import winreg

            with winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders",
            ) as key:
                value, _type = winreg.QueryValueEx(
                    key,
                    "{374DE290-123F-4565-9164-39C4925E467B}",
                )
            downloads = Path(os.path.expandvars(value))
            if downloads.exists():
                return downloads
        except Exception:
            pass
    for candidate in (Path.home() / "Downloads", Path.home() / "下载"):
        if candidate.exists():
            return candidate
    return Path.home()


def open_path(path: Path) -> None:
    try:
        if os.name == "nt":
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])
    except Exception:
        pass


def open_conversion_outputs(outputs: list[Path], output_dir: Path) -> None:
    existing = [path for path in outputs if path.exists()]
    if not existing:
        return
    folder = output_dir if output_dir.exists() else existing[0].parent
    open_path(folder)
    if len(existing) == 1:
        open_path(existing[0])


def collect_files(
    paths: Iterable[Path],
    recursive: bool = True,
    allowed_exts: set[str] | None = None,
) -> list[Path]:
    files: list[Path] = []
    allowed = allowed_exts or SUPPORTED_EXTS
    for path in paths:
        path = path.expanduser().resolve()
        if path.is_file():
            files.append(path)
        elif path.is_dir():
            pattern = "**/*" if recursive else "*"
            files.extend(p for p in path.glob(pattern) if p.is_file())
    return [p for p in files if p.suffix.lower() in allowed]


def normalized_image_frames(path: Path) -> list[Image.Image]:
    frames: list[Image.Image] = []
    with Image.open(path) as img:
        for frame in ImageSequence.Iterator(img):
            page = ImageOps.exif_transpose(frame.copy())
            if page.mode in ("RGBA", "LA") or (
                page.mode == "P" and "transparency" in page.info
            ):
                background = Image.new("RGB", page.size, "white")
                background.paste(page.convert("RGBA"), mask=page.convert("RGBA").split()[-1])
                page = background
            else:
                page = page.convert("RGB")
            frames.append(page)
    return frames


def image_pdf_page_size(image: Image.Image) -> tuple[float, float]:
    page_w = A4[0]
    img_w, img_h = image.size
    if img_w <= 0 or img_h <= 0:
        return A4
    return page_w, page_w * (img_h / img_w)


def draw_image_page(pdf: canvas.Canvas, image: Image.Image) -> None:
    page_w, page_h = image_pdf_page_size(image)
    pdf.setPageSize((page_w, page_h))
    pdf.drawImage(ImageReader(image), 0, 0, page_w, page_h, preserveAspectRatio=True)
    pdf.showPage()


def convert_image_to_pdf(source: Path, target: Path) -> None:
    frames = normalized_image_frames(source)
    pdf = canvas.Canvas(str(target), pagesize=A4)
    try:
        for frame in frames:
            draw_image_page(pdf, frame)
    finally:
        pdf.save()
        for frame in frames:
            frame.close()


def convert_images_to_single_pdf(images: list[Path], target: Path) -> None:
    pdf = canvas.Canvas(str(target), pagesize=A4)
    opened_frames: list[Image.Image] = []
    try:
        for image_path in images:
            frames = normalized_image_frames(image_path)
            opened_frames.extend(frames)
            for frame in frames:
                draw_image_page(pdf, frame)
    finally:
        pdf.save()
        for frame in opened_frames:
            frame.close()


def merge_pdfs(pdf_files: list[Path], target: Path) -> None:
    from pypdf import PdfReader, PdfWriter

    writer = PdfWriter()
    for pdf_file in pdf_files:
        reader = PdfReader(str(pdf_file))
        for page in reader.pages:
            writer.add_page(page)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("wb") as handle:
        writer.write(handle)


def normalize_pdf_width(source: Path, target_width: float = A4[0]) -> None:
    from pypdf import PdfReader, PdfWriter

    temp = source.with_name(f"{source.stem}_normalized_tmp{source.suffix}")
    reader = PdfReader(str(source))
    writer = PdfWriter()
    for page in reader.pages:
        width = float(page.mediabox.width)
        if width > 0 and abs(width - target_width) > 0.01:
            page.scale_by(target_width / width)
        writer.add_page(page)
    with temp.open("wb") as handle:
        writer.write(handle)
    temp.replace(source)


def normalize_page_ranges(ranges: str) -> str:
    replacements = {
        "\uFF0C": ",",
        "\u3001": ",",
        "\uFF1B": ",",
        ";": ",",
        "\uFF0D": "-",
        "\u2013": "-",
        "\u2014": "-",
        "~": "-",
        "\uFF5E": "-",
        "\u81F3": "-",
        "\u5230": "-",
        "\u311B": ",",
        "\u311C": "-",
        "\u301E": "-",
    }
    cleaned = ranges.strip()
    for old, new in replacements.items():
        cleaned = cleaned.replace(old, new)
    return cleaned

def parse_page_ranges(ranges: str, total_pages: int) -> list[int]:
    pages: list[int] = []
    seen: set[int] = set()
    cleaned = normalize_page_ranges(ranges)
    for part in cleaned.split(","):
        token = part.strip()
        if not token:
            continue
        if "-" in token:
            start_text, end_text = token.split("-", 1)
            start = int(start_text.strip())
            end = int(end_text.strip())
            if start > end:
                start, end = end, start
            numbers = range(start, end + 1)
        else:
            numbers = [int(token)]
        for page_number in numbers:
            if page_number < 1 or page_number > total_pages:
                raise ValueError(f"Page {page_number} is outside 1-{total_pages}")
            index = page_number - 1
            if index not in seen:
                pages.append(index)
                seen.add(index)
    if not pages:
        raise ValueError("No pages selected")
    return pages


def split_pdf_indices(source: Path, target: Path, page_indices: list[int]) -> int:
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(str(source))
    if not page_indices:
        raise ValueError("No pages selected")
    writer = PdfWriter()
    for index in page_indices:
        if index < 0 or index >= len(reader.pages):
            raise ValueError(f"Page {index + 1} is outside 1-{len(reader.pages)}")
        writer.add_page(reader.pages[index])
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("wb") as handle:
        writer.write(handle)
    return len(page_indices)


def trim_table(rows: list[list[object]]) -> list[list[str]]:
    occupied_rows: list[int] = []
    occupied_cols: set[int] = set()
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            if value not in (None, ""):
                occupied_rows.append(row_index)
                occupied_cols.add(col_index)
    if not occupied_rows or not occupied_cols:
        return []
    min_row, max_row = min(occupied_rows), max(occupied_rows)
    min_col, max_col = min(occupied_cols), max(occupied_cols)
    trimmed: list[list[str]] = []
    for row in rows[min_row : max_row + 1]:
        trimmed.append(
            ["" if value is None else str(value) for value in row[min_col : max_col + 1]]
        )
    return trimmed


def csv_used_table(source: Path) -> list[tuple[str, list[list[str]]]]:
    import csv

    rows: list[list[object]] = []
    with source.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        rows.extend(list(csv.reader(handle)))
    table = trim_table(rows)
    return [(source.stem, table)] if table else []


def workbook_used_tables(source: Path) -> list[tuple[str, list[list[str]]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(source, data_only=True, read_only=True)
    tables: list[tuple[str, list[list[str]]]] = []
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            table = trim_table(rows)
            if table:
                tables.append((sheet.title, table))
    finally:
        workbook.close()
    return tables


def convert_spreadsheet_to_pdf(source: Path, target: Path) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.pdfmetrics import registerFont
    from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer
    from reportlab.platypus import TableStyle

    registerFont(UnicodeCIDFont("STSong-Light"))
    tables = csv_used_table(source) if source.suffix.lower() == ".csv" else workbook_used_tables(source)
    if not tables:
        raise RuntimeError("No table data found")

    page_size = landscape(A4)
    margin = 10 * mm
    usable_width = page_size[0] - margin * 2
    styles = getSampleStyleSheet()
    styles["Normal"].fontName = "STSong-Light"
    styles["Title"].fontName = "STSong-Light"
    styles["Normal"].fontSize = 8

    story: list[object] = []
    for sheet_index, (sheet_name, table) in enumerate(tables):
        if sheet_index:
            story.append(PageBreak())
        story.append(Paragraph(sheet_name, styles["Title"]))
        story.append(Spacer(1, 4 * mm))
        col_count = max(len(row) for row in table)
        normalized = [row + [""] * (col_count - len(row)) for row in table]
        col_width = min(36 * mm, usable_width / max(1, col_count))
        pdf_table = LongTable(
            normalized,
            colWidths=[col_width] * col_count,
            repeatRows=1 if len(normalized) > 1 else 0,
        )
        pdf_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f3f7")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b8c0cc")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(pdf_table)

    doc = SimpleDocTemplate(
        str(target),
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
    )
    doc.build(story)


def convert_text_to_pdf(source: Path, target: Path) -> None:
    from reportlab.lib.units import mm
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.pdfmetrics import registerFont

    registerFont(UnicodeCIDFont("STSong-Light"))

    page_w, page_h = A4
    margin = 18 * mm
    line_height = 14
    font_name = "STSong-Light"
    font_size = 10
    max_chars = 86

    text = source.read_text(encoding="utf-8", errors="replace")
    pdf = canvas.Canvas(str(target), pagesize=A4)
    pdf.setTitle(source.stem)
    pdf.setFont(font_name, font_size)

    y = page_h - margin
    for raw_line in text.splitlines() or [""]:
        line = raw_line.replace("\t", "    ")
        chunks = [line[i : i + max_chars] for i in range(0, len(line), max_chars)] or [""]
        for chunk in chunks:
            if y < margin:
                pdf.showPage()
                pdf.setFont(font_name, font_size)
                y = page_h - margin
            pdf.drawString(margin, y, chunk)
            y -= line_height
    pdf.save()


def find_libreoffice() -> str | None:
    candidates = [
        "soffice",
        "libreoffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for candidate in candidates:
        found = shutil.which(candidate)
        if found:
            return found
        if Path(candidate).exists():
            return candidate
    return None


def check_shutdown_requested() -> None:
    if SHUTDOWN_EVENT.is_set():
        raise RuntimeError("Conversion cancelled")


def taskkill_pid(pid: int) -> None:
    if pid <= 0:
        return
    if os.name == "nt":
        flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        try:
            subprocess.run(
                ["taskkill", "/PID", str(pid), "/T", "/F"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=flags,
                timeout=5,
            )
        except Exception:
            pass


def terminate_process(proc: subprocess.Popen) -> None:
    if proc.poll() is not None:
        return
    if os.name == "nt":
        taskkill_pid(proc.pid)
        return
    try:
        proc.terminate()
        proc.wait(timeout=3)
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


def register_subprocess(proc: subprocess.Popen) -> None:
    with ACTIVE_PROCESS_LOCK:
        ACTIVE_SUBPROCESSES.add(proc)


def unregister_subprocess(proc: subprocess.Popen) -> None:
    with ACTIVE_PROCESS_LOCK:
        ACTIVE_SUBPROCESSES.discard(proc)


def register_external_pids(pids: set[int]) -> None:
    if not pids:
        return
    with ACTIVE_PROCESS_LOCK:
        ACTIVE_EXTERNAL_PIDS.update(pid for pid in pids if pid > 0)


def unregister_external_pids(pids: set[int]) -> None:
    if not pids:
        return
    with ACTIVE_PROCESS_LOCK:
        ACTIVE_EXTERNAL_PIDS.difference_update(pids)


def terminate_tracked_processes() -> None:
    with ACTIVE_PROCESS_LOCK:
        subprocesses = list(ACTIVE_SUBPROCESSES)
        external_pids = list(ACTIVE_EXTERNAL_PIDS)
    for proc in subprocesses:
        terminate_process(proc)
    for pid in external_pids:
        taskkill_pid(pid)


def office_process_snapshot() -> set[int]:
    if os.name != "nt":
        return set()
    flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    try:
        completed = subprocess.run(
            ["tasklist", "/FO", "CSV", "/NH"],
            capture_output=True,
            text=True,
            creationflags=flags,
            timeout=5,
        )
    except Exception:
        return set()
    pids: set[int] = set()
    for row in csv.reader(completed.stdout.splitlines()):
        if len(row) < 2:
            continue
        if row[0].strip().upper() not in OFFICE_PROCESS_NAMES:
            continue
        try:
            pids.add(int(row[1]))
        except ValueError:
            pass
    return pids


def wait_for_tracked_process(
    proc: subprocess.Popen,
    timeout_seconds: int,
) -> tuple[int, str, str]:
    deadline = time.monotonic() + timeout_seconds
    while True:
        if SHUTDOWN_EVENT.is_set():
            terminate_process(proc)
            raise RuntimeError("Conversion cancelled")
        remaining = max(0.1, min(0.25, deadline - time.monotonic()))
        try:
            stdout, stderr = proc.communicate(timeout=remaining)
            return proc.returncode or 0, stdout or "", stderr or ""
        except subprocess.TimeoutExpired:
            if time.monotonic() >= deadline:
                terminate_process(proc)
                raise RuntimeError("Conversion timed out")


def convert_with_libreoffice(source: Path, output_dir: Path) -> Path:
    soffice = find_libreoffice()
    if not soffice:
        raise RuntimeError("LibreOffice was not found")

    check_shutdown_requested()
    before = set(output_dir.glob("*.pdf"))
    flags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
    proc = subprocess.Popen(
        [
            soffice,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(output_dir),
            str(source),
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        creationflags=flags,
    )
    register_subprocess(proc)
    try:
        returncode, stdout, stderr = wait_for_tracked_process(proc, 180)
    finally:
        unregister_subprocess(proc)

    after = set(output_dir.glob("*.pdf"))
    created = sorted(after - before, key=lambda p: p.stat().st_mtime, reverse=True)
    expected = output_dir / f"{source.stem}.pdf"

    if returncode != 0:
        detail = (stderr or stdout or "").strip()
        raise RuntimeError(detail or "LibreOffice conversion failed")
    if expected.exists():
        return expected
    if created:
        return created[0]
    raise RuntimeError("LibreOffice did not create a PDF file")


def dispatch_com_application(win32_client, progids: tuple[str, ...]):
    errors: list[str] = []
    for progid in progids:
        try:
            return win32_client.DispatchEx(progid), progid
        except Exception as exc:  # noqa: BLE001 - each provider is a fallback.
            errors.append(f"{progid}: {exc}")
    raise RuntimeError("No compatible Microsoft Office or WPS COM app found. " + " | ".join(errors))


def safe_com_call(obj, method_name: str, *args) -> None:
    if obj is None:
        return
    try:
        getattr(obj, method_name)(*args)
    except Exception:
        pass


def close_com_document(document) -> None:
    if document is None:
        return
    try:
        document.Close(False)
    except Exception:
        safe_com_call(document, "Close")


def set_visible_false(app) -> None:
    try:
        app.Visible = False
    except Exception:
        pass


def open_presentation(app, source: Path):
    try:
        return app.Presentations.Open(str(source), WithWindow=False)
    except Exception:
        return app.Presentations.Open(str(source), False, False, False)


def export_word_pdf(doc, target: Path) -> None:
    try:
        doc.ExportAsFixedFormat(str(target), 17)
    except Exception:
        doc.SaveAs(str(target), FileFormat=17)


def export_workbook_pdf(workbook, target: Path) -> None:
    try:
        workbook.ExportAsFixedFormat(0, str(target))
    except Exception:
        workbook.SaveAs(str(target), FileFormat=57)


def convert_with_office_com(source: Path, target: Path) -> None:
    if os.name != "nt":
        raise RuntimeError("Microsoft Office or WPS automation is only available on Windows")

    check_shutdown_requested()
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError("pywin32 is missing, so Microsoft Office/WPS cannot be used") from exc

    ext = source.suffix.lower()
    with COM_CONVERSION_LOCK:
        check_shutdown_requested()
        pythoncom.CoInitialize()
        app = None
        document = None
        known_pids = office_process_snapshot()
        tracked_pids: set[int] = set()

        def track_new_office_processes() -> None:
            new_pids = office_process_snapshot() - known_pids - tracked_pids
            if new_pids:
                tracked_pids.update(new_pids)
                register_external_pids(new_pids)

        try:
            if ext in {".doc", ".docx", ".rtf"}:
                app, _provider = dispatch_com_application(
                    win32com.client,
                    ("Word.Application", "KWPS.Application", "WPS.Application"),
                )
                track_new_office_processes()
                check_shutdown_requested()
                set_visible_false(app)
                document = app.Documents.Open(str(source))
                track_new_office_processes()
                check_shutdown_requested()
                export_word_pdf(document, target)
            elif ext in {".ppt", ".pptx"}:
                app, _provider = dispatch_com_application(
                    win32com.client,
                    ("PowerPoint.Application", "KWPP.Application"),
                )
                track_new_office_processes()
                check_shutdown_requested()
                set_visible_false(app)
                document = open_presentation(app, source)
                track_new_office_processes()
                check_shutdown_requested()
                document.SaveAs(str(target), 32)
            elif ext in {".xls", ".xlsx", ".xlsm", ".csv"}:
                app, _provider = dispatch_com_application(
                    win32com.client,
                    ("Excel.Application", "KET.Application", "ET.Application"),
                )
                track_new_office_processes()
                check_shutdown_requested()
                set_visible_false(app)
                document = app.Workbooks.Open(str(source))
                track_new_office_processes()
                check_shutdown_requested()
                export_workbook_pdf(document, target)
            else:
                raise RuntimeError(f"Microsoft Office/WPS does not support this extension: {ext}")
        finally:
            if document is not None:
                close_com_document(document)
            if app is not None:
                safe_com_call(app, "Quit")
            if SHUTDOWN_EVENT.is_set():
                terminate_tracked_processes()
            unregister_external_pids(tracked_pids)
            pythoncom.CoUninitialize()


def convert_office_to_pdf(source: Path, target: Path, output_dir: Path) -> Path:
    if find_libreoffice():
        created = convert_with_libreoffice(source, output_dir)
        if created != target:
            if target.exists():
                target.unlink()
            created.rename(target)
        if source.suffix.lower() in {".ppt", ".pptx", ".odp"}:
            normalize_pdf_width(target)
        return target

    convert_with_office_com(source, target)
    if source.suffix.lower() in {".ppt", ".pptx", ".odp"}:
        normalize_pdf_width(target)
    return target


def convert_one(source: Path, output_dir: Path) -> ConvertResult:
    check_shutdown_requested()
    output_dir.mkdir(parents=True, exist_ok=True)
    ext = source.suffix.lower()
    target = safe_pdf_name(source, output_dir)

    try:
        if ext in IMAGE_EXTS:
            convert_image_to_pdf(source, target)
        elif ext in TEXT_EXTS:
            convert_text_to_pdf(source, target)
        elif ext in SPREADSHEET_TABLE_EXTS:
            convert_spreadsheet_to_pdf(source, target)
        elif ext in OFFICE_EXTS:
            convert_office_to_pdf(source, target, output_dir)
        elif ext == ".pdf":
            shutil.copy2(source, target)
        else:
            return ConvertResult(source, False, "Unsupported file type")
        return ConvertResult(source, True, "Done", target)
    except Exception as exc:  # noqa: BLE001 - GUI needs friendly per-file failures.
        return ConvertResult(source, False, str(exc), None)


def publish_results(
    results: list[ConvertResult],
    on_result: Callable[[ConvertResult], None] | None = None,
) -> list[ConvertResult]:
    if on_result:
        for result in results:
            on_result(result)
    return results


def discarded_batch_results(staged: list[ConvertResult], reason: str) -> list[ConvertResult]:
    if not staged:
        return [ConvertResult(source=Path("batch"), ok=False, message=reason, output=None)]
    discarded: list[ConvertResult] = []
    for result in staged:
        message = result.message if not result.ok else reason
        discarded.append(ConvertResult(result.source, False, message, None))
    return discarded


def publish_discarded_results(
    staged: list[ConvertResult],
    reason: str,
    on_result: Callable[[ConvertResult], None] | None = None,
) -> list[ConvertResult]:
    return publish_results(discarded_batch_results(staged, reason), on_result)


def move_staged_pdf(staged_pdf: Path, final_target: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    if final_target.exists():
        final_target = safe_pdf_name(final_target, output_dir)
    shutil.move(str(staged_pdf), str(final_target))
    return final_target


def publish_staged_outputs(
    staged_outputs: list[StagedOutput],
    output_dir: Path,
) -> list[ConvertResult]:
    moved_outputs: list[Path] = []
    final_results: list[ConvertResult] = []
    try:
        for staged in staged_outputs:
            moved = move_staged_pdf(staged.staged_pdf, staged.final_target, output_dir)
            moved_outputs.append(moved)
            final_results.append(
                ConvertResult(staged.result.source, True, staged.result.message, moved)
            )
    except Exception:
        for moved in moved_outputs:
            try:
                moved.unlink()
            except OSError:
                pass
        raise
    return final_results


def batch_convert(
    inputs: list[Path],
    output_dir: Path,
    recursive: bool = True,
    allowed_exts: set[str] | None = None,
    merge_images: bool = False,
    merge_all: bool = False,
    merged_name: str = "merged_images.pdf",
    on_result: Callable[[ConvertResult], None] | None = None,
) -> list[ConvertResult]:
    files = collect_files(inputs, recursive=recursive, allowed_exts=allowed_exts)

    if not files:
        return publish_results([], on_result)

    if merge_all:
        temp_dir = Path(tempfile.mkdtemp(prefix="pdf_converter_"))
        converted: list[Path] = []
        staged_results: list[ConvertResult] = []
        try:
            for file in files:
                if SHUTDOWN_EVENT.is_set():
                    break
                result = convert_one(file, temp_dir)
                staged_results.append(result)
                if not result.ok:
                    break
                if result.ok and result.output:
                    converted.append(result.output)
            if SHUTDOWN_EVENT.is_set():
                return publish_discarded_results(
                    staged_results,
                    "Conversion cancelled; generated files were removed",
                    on_result,
                )
            if len(staged_results) != len(files) or not all(item.ok for item in staged_results):
                return publish_discarded_results(
                    staged_results,
                    "Batch failed; generated files were removed",
                    on_result,
                )
            if not converted:
                return publish_results(
                    [
                        ConvertResult(
                            Path(merged_name),
                            False,
                            "No PDF files were created for merging",
                            None,
                        )
                    ],
                    on_result,
                )

            staged_merged = safe_pdf_name(Path(merged_name), temp_dir)
            merge_pdfs(converted, staged_merged)
            check_shutdown_requested()
            final_target = safe_pdf_name(Path(merged_name), output_dir)
            moved = move_staged_pdf(staged_merged, final_target, output_dir)
            results: list[ConvertResult] = [
                ConvertResult(item.source, True, "Prepared for merge", None)
                for item in staged_results
            ]
            results.append(
                ConvertResult(
                    source=moved,
                    ok=True,
                    message=f"Merged {len(converted)} PDF files",
                    output=moved,
                )
            )
            return publish_results(results, on_result)
        except Exception as exc:  # noqa: BLE001 - clean partial batch output.
            return publish_discarded_results(
                staged_results,
                f"Batch failed; generated files were removed: {exc}",
                on_result,
            )
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    staged_results: list[ConvertResult] = []
    staged_outputs: list[StagedOutput] = []
    temp_dir = Path(tempfile.mkdtemp(prefix="pdf_converter_"))
    expected_stage_count = len(files)
    if merge_images:
        image_files = [p for p in files if p.suffix.lower() in IMAGE_EXTS]
        other_files = [p for p in files if p.suffix.lower() not in IMAGE_EXTS]
        expected_stage_count = len(other_files) + (1 if image_files else 0)
        if image_files:
            staged_target = safe_pdf_name(Path(merged_name), temp_dir)
            try:
                convert_images_to_single_pdf(image_files, staged_target)
                result = ConvertResult(
                    source=Path(merged_name),
                    ok=True,
                    message=f"Merged {len(image_files)} images",
                    output=staged_target,
                )
            except Exception as exc:  # noqa: BLE001
                result = ConvertResult(
                    source=Path(merged_name),
                    ok=False,
                    message=f"Image merge failed: {exc}",
                    output=None,
                )
            staged_results.append(result)
            if result.ok and result.output:
                staged_outputs.append(
                    StagedOutput(result, result.output, output_dir / staged_target.name)
                )
        files = other_files

    try:
        for file in files:
            if SHUTDOWN_EVENT.is_set():
                break
            result = convert_one(file, temp_dir)
            staged_results.append(result)
            if not result.ok:
                break
            if result.output:
                staged_outputs.append(
                    StagedOutput(result, result.output, safe_pdf_name(result.source, output_dir))
                )

        if SHUTDOWN_EVENT.is_set():
            return publish_discarded_results(
                staged_results,
                "Conversion cancelled; generated files were removed",
                on_result,
            )
        if len(staged_results) < expected_stage_count:
            return publish_discarded_results(
                staged_results,
                "Batch did not complete; generated files were removed",
                on_result,
            )
        if not all(item.ok for item in staged_results):
            return publish_discarded_results(
                staged_results,
                "Batch failed; generated files were removed",
                on_result,
            )

        return publish_results(publish_staged_outputs(staged_outputs, output_dir), on_result)
    except Exception as exc:  # noqa: BLE001 - clean partial batch output.
        return publish_discarded_results(
            staged_results,
            f"Batch failed; generated files were removed: {exc}",
            on_result,
        )
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def run_gui() -> None:
    enable_high_dpi()
    import tkinter as tk
    import tkinter.font as tkfont
    from tkinter import filedialog, messagebox, ttk

    try:
        from tkinterdnd2 import DND_FILES, TkinterDnD
    except ImportError:
        DND_FILES = None
        TkinterDnD = None

    root = TkinterDnD.Tk() if TkinterDnD else tk.Tk()
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    window_w = min(max(1180, int(screen_w * 0.72)), max(900, screen_w - 80))
    window_h = min(max(780, int(screen_h * 0.78)), max(640, screen_h - 80))
    pos_x = max(0, (screen_w - window_w) // 2)
    pos_y = max(0, (screen_h - window_h) // 2)
    root.geometry(f"{window_w}x{window_h}+{pos_x}+{pos_y}")
    root.minsize(min(960, max(760, screen_w - 80)), min(620, max(520, screen_h - 80)))
    try:
        root.tk.call("tk", "scaling", max(1.2, root.winfo_fpixels("1i") / 72))
    except tk.TclError:
        pass

    default_font = tkfont.nametofont("TkDefaultFont")
    default_font.configure(family="Microsoft YaHei UI", size=10)
    text_font = tkfont.nametofont("TkTextFont")
    text_font.configure(family="Microsoft YaHei UI", size=10)
    root.option_add("*Font", default_font)

    style = ttk.Style(root)
    if "vista" in style.theme_names():
        style.theme_use("vista")
    style.configure("TButton", padding=(10, 7))
    style.configure("TCheckbutton", padding=(4, 4))
    style.configure("TLabel", padding=(2, 2))

    selected: list[Path] = []
    default_output = default_output_dir()
    settings = load_settings()

    recursive_var = tk.BooleanVar(value=True)
    output_var = tk.StringVar(value=str(default_output))
    language_code_var = tk.StringVar(value=settings["language"])
    language_display_var = tk.StringVar(value=LANGUAGES[settings["language"]])
    type_var = tk.StringVar(value=PROFILE_LABELS[settings["language"]]["All supported files"])
    merged_name_var = tk.StringVar(value="converted.pdf")
    status_var = tk.StringVar(value=TEXT[settings["language"]]["drop_hint"])
    closing_var = tk.BooleanVar(value=False)
    active_workers: list[threading.Thread] = []
    close_deadline = {"time": 0.0}

    root.columnconfigure(0, weight=1)
    root.rowconfigure(3, weight=1)

    top = ttk.Frame(root, padding=14)
    top.grid(row=0, column=0, sticky="ew")
    top.columnconfigure(4, weight=1)

    def current_language() -> str:
        return language_code_var.get()

    def tr(key: str, **kwargs: object) -> str:
        text = TEXT[current_language()][key]
        return text.format(**kwargs) if kwargs else text

    def profile_display(profile_key: str) -> str:
        return PROFILE_LABELS[current_language()][profile_key]

    def profile_key_from_display(display: str) -> str:
        for labels in PROFILE_LABELS.values():
            for key, value in labels.items():
                if value == display:
                    return key
        return "All supported files"

    def refresh_list() -> None:
        file_list.delete(0, tk.END)
        for i, path in enumerate(selected, start=1):
            file_list.insert(tk.END, f"{i:02d}. {path}")
        status_var.set(tr("selected", count=len(selected)))

    def add_paths(paths: Iterable[Path]) -> None:
        existing = {p.resolve() for p in selected if p.exists()}
        allowed = TYPE_PROFILES.get(profile_key_from_display(type_var.get()), SUPPORTED_EXTS)
        rejected: list[Path] = []
        added_count = 0
        for path in paths:
            resolved = path.expanduser().resolve()
            if resolved.is_dir():
                candidates = collect_files([resolved], recursive=recursive_var.get(), allowed_exts=allowed)
                if not candidates:
                    rejected.append(resolved)
            else:
                if resolved.suffix.lower() in allowed:
                    candidates = [resolved]
                else:
                    candidates = []
                    rejected.append(resolved)
            for candidate in candidates:
                if candidate not in existing:
                    selected.append(candidate)
                    existing.add(candidate)
                    added_count += 1
        refresh_list()
        if rejected:
            for path in rejected[:20]:
                append_log(tr("add_failed_log", path=path))
            if len(rejected) > 20:
                append_log(tr("add_failed_log", path=f"... +{len(rejected) - 20}"))
            messagebox.showwarning(
                tr("add_failed_title"),
                tr("add_failed_body", count=len(rejected)),
            )
        elif added_count == 0:
            append_log(tr("add_none"))

    def add_files() -> None:
        profile_key = profile_key_from_display(type_var.get())
        profile = profile_display(profile_key)
        exts = TYPE_PROFILES.get(profile_key, SUPPORTED_EXTS)
        pattern = " ".join(f"*{ext}" for ext in sorted(exts))
        files = filedialog.askopenfilenames(
            title=tr("choose_files"),
            filetypes=[
                (profile, pattern),
                (profile_display("All supported files"), " ".join(f"*{ext}" for ext in sorted(SUPPORTED_EXTS))),
                (tr("all_files"), "*.*"),
            ],
        )
        add_paths(Path(f) for f in files)

    def add_folder() -> None:
        folder = filedialog.askdirectory(title=tr("choose_folder"))
        if folder:
            add_paths([Path(folder)])

    def clear_items() -> None:
        selected.clear()
        refresh_list()

    def choose_output() -> None:
        folder = filedialog.askdirectory(title=tr("choose_output"))
        if folder:
            output_var.set(folder)

    add_files_button = ttk.Button(top, command=add_files)
    add_files_button.grid(row=0, column=0, padx=(0, 8))
    add_folder_button = ttk.Button(top, command=add_folder)
    add_folder_button.grid(row=0, column=1, padx=(0, 8))
    clear_button = ttk.Button(top, command=clear_items)
    clear_button.grid(row=0, column=2, padx=(0, 12))
    recursive_check = ttk.Checkbutton(top, variable=recursive_var)
    recursive_check.grid(row=0, column=3, padx=(0, 12))
    language_label = ttk.Label(top)
    language_label.grid(row=0, column=5, sticky="e", padx=(12, 8))
    language_menu = ttk.Combobox(
        top,
        textvariable=language_display_var,
        values=list(LANGUAGES.values()),
        state="readonly",
        width=10,
    )
    language_menu.grid(row=0, column=6, sticky="e")

    options = ttk.Frame(root, padding=(14, 0, 14, 10))
    options.grid(row=1, column=0, sticky="ew")
    options.columnconfigure(1, weight=1)
    options.columnconfigure(3, weight=1)
    convert_type_label = ttk.Label(options)
    convert_type_label.grid(row=0, column=0, sticky="w", padx=(0, 8))
    type_menu = ttk.Combobox(
        options,
        textvariable=type_var,
        state="readonly",
    )
    type_menu.grid(row=0, column=1, sticky="ew", padx=(0, 12))
    merged_name_label = ttk.Label(options)
    merged_name_label.grid(row=0, column=2, sticky="w", padx=(0, 8))
    ttk.Entry(options, textvariable=merged_name_var).grid(row=0, column=3, sticky="ew")

    output_frame = ttk.Frame(root, padding=(14, 0, 14, 10))
    output_frame.columnconfigure(1, weight=1)
    output_frame.grid(row=2, column=0, sticky="ew")
    output_label = ttk.Label(output_frame)
    output_label.grid(row=0, column=0, padx=(0, 8))
    ttk.Entry(output_frame, textvariable=output_var).grid(row=0, column=1, sticky="ew", padx=(0, 8))
    browse_button = ttk.Button(output_frame, command=choose_output)
    browse_button.grid(row=0, column=2)

    main = ttk.Frame(root, padding=(14, 0, 14, 10))
    main.grid(row=3, column=0, sticky="nsew")
    main.columnconfigure(0, weight=1)
    main.columnconfigure(2, weight=0)
    main.rowconfigure(0, weight=8, minsize=300)
    main.rowconfigure(1, weight=1, minsize=70)

    file_list = tk.Listbox(
        main,
        activestyle="none",
        selectmode=tk.EXTENDED,
        font=("Microsoft YaHei UI", 10),
        highlightthickness=1,
        relief="solid",
    )
    file_list.grid(row=0, column=0, sticky="nsew")
    scrollbar = ttk.Scrollbar(main, orient="vertical", command=file_list.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")
    file_list.configure(yscrollcommand=scrollbar.set)

    side = ttk.Frame(main, padding=(10, 0, 0, 0))
    side.grid(row=0, column=2, sticky="ns")

    def selected_indices() -> list[int]:
        return list(file_list.curselection())

    def move_selection(delta: int) -> None:
        indices = selected_indices()
        if not indices:
            return
        moving = indices if delta < 0 else list(reversed(indices))
        for idx in moving:
            new_idx = idx + delta
            if 0 <= new_idx < len(selected):
                selected[idx], selected[new_idx] = selected[new_idx], selected[idx]
        refresh_list()
        for idx in [i + delta for i in indices if 0 <= i + delta < len(selected)]:
            file_list.selection_set(idx)

    def remove_selection() -> None:
        for idx in reversed(selected_indices()):
            selected.pop(idx)
        refresh_list()

    def choose_pdf_pages_with_preview(source: Path) -> list[int] | None:
        try:
            import fitz  # type: ignore
            from PIL import ImageTk

            doc = fitz.open(str(source))
        except Exception:
            messagebox.showwarning(tr("split_pdf"), tr("split_preview_failed"))
            return None

        selected_pages: set[int] = set()
        result: list[int] | None = None
        thumbnails: list[object] = []
        done_var = tk.BooleanVar(value=False)
        range_var = tk.StringVar(value="")
        previous_title = root.title()
        root.title(tr("split_pdf"))

        managed_widgets = list(root.grid_slaves())
        for widget in managed_widgets:
            widget.grid_remove()

        split_frame = ttk.Frame(root, padding=14)
        split_frame.grid(row=0, column=0, rowspan=5, sticky="nsew")
        split_frame.columnconfigure(0, weight=1)
        split_frame.rowconfigure(1, weight=1)

        toolbar = ttk.Frame(split_frame)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        toolbar.columnconfigure(4, weight=1)

        canvas_area = ttk.Frame(split_frame)
        canvas_area.grid(row=1, column=0, sticky="nsew")
        canvas_area.columnconfigure(0, weight=1)
        canvas_area.rowconfigure(0, weight=1)
        preview_canvas = tk.Canvas(canvas_area, highlightthickness=0, background="#f4f6f8")
        preview_canvas.grid(row=0, column=0, sticky="nsew")
        preview_scroll = ttk.Scrollbar(canvas_area, orient="vertical", command=preview_canvas.yview)
        preview_scroll.grid(row=0, column=1, sticky="ns")
        preview_canvas.configure(yscrollcommand=preview_scroll.set)

        layout_state = {"columns": 0, "tile_width": 0}
        thumbnail_zoom = {"columns": None}
        resize_job = {"id": None}

        def schedule_render(delay_ms: int = 120) -> None:
            if resize_job["id"] is not None:
                try:
                    root.after_cancel(resize_job["id"])
                except Exception:
                    pass

            def run_render() -> None:
                resize_job["id"] = None
                render_grid()

            resize_job["id"] = root.after(delay_ms, run_render)

        def update_canvas_width(event) -> None:  # noqa: ANN001 - tkinter event object.
            schedule_render()

        preview_canvas.bind("<Configure>", update_canvas_width)

        def zoom_thumbnails(event) -> str:  # noqa: ANN001 - tkinter event object.
            delta = getattr(event, "delta", 0)
            if not delta:
                return "break"
            direction = 1 if delta > 0 else -1
            steps = max(1, min(4, abs(delta) // 120 if abs(delta) >= 120 else 1))
            available_width = max(260, preview_canvas.winfo_width() - 28)
            gap_x = 22
            min_tile_width = 130
            default_tile_width = 260
            max_columns = max(1, min(len(doc), int((available_width + gap_x) // (min_tile_width + gap_x))))
            auto_columns = max(1, min(max_columns, int((available_width + gap_x) // (default_tile_width + gap_x))))
            current_columns = thumbnail_zoom["columns"] or layout_state["columns"] or auto_columns
            current_columns = max(1, min(max_columns, current_columns))
            new_columns = max(1, min(max_columns, current_columns - direction * steps))
            if new_columns == current_columns:
                return "break"
            thumbnail_zoom["columns"] = new_columns
            layout_state["columns"] = 0
            layout_state["tile_width"] = 0
            schedule_render(70)
            return "break"

        def on_mousewheel(event) -> str:  # noqa: ANN001 - tkinter event object.
            if getattr(event, "state", 0) & 0x0004:
                return zoom_thumbnails(event)
            scrollregion = preview_canvas.cget("scrollregion")
            total_height = 0
            if scrollregion:
                coords = [float(part) for part in scrollregion.split()]
                total_height = coords[3] - coords[1]
            if total_height <= preview_canvas.winfo_height():
                preview_canvas.yview_moveto(0)
                return "break"
            if event.delta:
                preview_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        def bind_mousewheel(_event=None) -> None:  # noqa: ANN001 - tkinter callback.
            preview_canvas.bind_all("<MouseWheel>", on_mousewheel)

        def unbind_mousewheel(_event=None) -> None:  # noqa: ANN001 - tkinter callback.
            preview_canvas.unbind_all("<MouseWheel>")

        canvas_area.bind("<Enter>", bind_mousewheel)
        canvas_area.bind("<Leave>", unbind_mousewheel)

        range_sync = {"active": False}
        placeholder_visible = {"value": False}

        def format_page_ranges(page_indices: list[int]) -> str:
            if not page_indices:
                return ""
            parts: list[str] = []
            start = previous = page_indices[0] + 1
            for page_index in page_indices[1:]:
                page_number = page_index + 1
                if page_number == previous + 1:
                    previous = page_number
                    continue
                parts.append(str(start) if start == previous else f"{start}-{previous}")
                start = previous = page_number
            parts.append(str(start) if start == previous else f"{start}-{previous}")
            return ",".join(parts)

        def set_range_text_from_selection() -> None:
            text = format_page_ranges(sorted(selected_pages))
            range_sync["active"] = True
            try:
                placeholder_visible["value"] = False
                range_entry.configure(style="TEntry")
                range_var.set(text)
            finally:
                range_sync["active"] = False
            if not text:
                show_range_placeholder()

        def apply_range_text_to_selection(show_errors: bool = False) -> bool:
            if range_sync["active"]:
                return True
            ranges = "" if placeholder_visible["value"] else range_var.get().strip()
            if not ranges:
                selected_pages.clear()
                for variable in page_vars.values():
                    variable.set(False)
                return True
            try:
                page_indices = parse_page_ranges(ranges, len(doc))
            except Exception as exc:  # noqa: BLE001 - page range typing can be temporarily incomplete.
                if show_errors:
                    messagebox.showerror(tr("split_pdf"), tr("split_failed", error=exc))
                return False
            selected_pages.clear()
            selected_pages.update(page_indices)
            for index, variable in page_vars.items():
                variable.set(index in selected_pages)
            return True

        def set_all(value: bool) -> None:
            if value:
                selected_pages.update(range(len(doc)))
            else:
                selected_pages.clear()
            for index, variable in page_vars.items():
                variable.set(index in selected_pages)
            set_range_text_from_selection()

        def finish_selection() -> None:
            nonlocal result
            if not apply_range_text_to_selection(True):
                return
            if not selected_pages:
                messagebox.showwarning(tr("split_pdf"), tr("split_no_pages_selected"))
                return
            result = sorted(selected_pages)
            done_var.set(True)

        def back() -> None:
            done_var.set(True)

        back_button = ttk.Button(toolbar, text=tr("back"), command=back)
        back_button.grid(row=0, column=0, sticky="w")
        select_all_button = ttk.Button(toolbar, text=tr("select_all"), command=lambda: set_all(True))
        select_all_button.grid(row=0, column=1, padx=(10, 0))
        clear_selection_button = ttk.Button(toolbar, text=tr("clear_selection"), command=lambda: set_all(False))
        clear_selection_button.grid(row=0, column=2, sticky="w", padx=(8, 0))
        range_label = ttk.Label(toolbar, text=tr("split_range_label"))
        range_label.grid(row=0, column=3, sticky="e", padx=(14, 6))
        range_box = ttk.Frame(toolbar)
        range_box.grid(row=0, column=4, sticky="ew")
        range_box.columnconfigure(0, weight=1)
        range_entry = ttk.Entry(range_box, textvariable=range_var, width=36)
        range_entry.grid(row=0, column=0, sticky="ew")
        range_style = ttk.Style(range_entry)
        range_style.configure("Placeholder.TEntry", foreground="#8a8f98")

        def show_range_placeholder() -> None:
            if range_var.get().strip():
                return
            range_sync["active"] = True
            try:
                placeholder_visible["value"] = True
                range_entry.configure(style="Placeholder.TEntry")
                range_var.set(tr("split_range_placeholder"))
            finally:
                range_sync["active"] = False

        def hide_range_placeholder() -> None:
            if not placeholder_visible["value"]:
                return
            range_sync["active"] = True
            try:
                placeholder_visible["value"] = False
                range_entry.configure(style="TEntry")
                range_var.set("")
            finally:
                range_sync["active"] = False

        range_entry.bind("<FocusIn>", lambda _event: hide_range_placeholder())
        range_entry.bind("<FocusOut>", lambda _event: show_range_placeholder())
        show_range_placeholder()
        export_selected_button = ttk.Button(toolbar, text=tr("export_selected"), command=finish_selection)
        export_selected_button.grid(row=0, column=5, sticky="e", padx=(10, 0))
        loading_label = ttk.Label(toolbar, text=tr("split_loading"))
        loading_label.grid(row=0, column=1, columnspan=8, sticky="w", padx=(12, 0))
        action_widgets = [
            select_all_button,
            clear_selection_button,
            range_label,
            range_box,
            export_selected_button,
        ]

        def set_preview_loading(is_loading: bool) -> None:
            if is_loading:
                for widget in action_widgets:
                    widget.grid_remove()
                loading_label.grid()
            else:
                loading_label.grid_remove()
                for widget in action_widgets:
                    widget.grid()

        page_vars: dict[int, tk.BooleanVar] = {
            index: tk.BooleanVar(value=False) for index in range(len(doc))
        }
        checkbox_widgets: list[object] = []
        page_bounds: dict[int, tuple[int, int, int, int]] = {}
        render_state = {"token": 0, "after": None, "rendering": False}
        drag_select = {"start": None, "rect": None, "moved": False}

        def on_range_changed(*_args) -> None:  # noqa: ANN002 - tkinter trace passes variable details.
            if placeholder_visible["value"]:
                return
            apply_range_text_to_selection(False)

        range_var.trace_add("write", on_range_changed)

        def sync_selection(page_index: int, variable: tk.BooleanVar, user_action: bool = False) -> None:
            if variable.get():
                selected_pages.add(page_index)
            else:
                selected_pages.discard(page_index)
            if user_action:
                set_range_text_from_selection()

        def toggle_from_page(page_index: int, variable: tk.BooleanVar) -> None:
            variable.set(not variable.get())
            sync_selection(page_index, variable)
            set_range_text_from_selection()

        def page_at(canvas_x: float, canvas_y: float) -> int | None:
            for page_index, (x1, y1, x2, y2) in page_bounds.items():
                if x1 <= canvas_x <= x2 and y1 <= canvas_y <= y2:
                    return page_index
            return None

        def rectangles_intersect(
            first: tuple[float, float, float, float],
            second: tuple[int, int, int, int],
        ) -> bool:
            ax1, ay1, ax2, ay2 = first
            bx1, by1, bx2, by2 = second
            return ax1 <= bx2 and ax2 >= bx1 and ay1 <= by2 and ay2 >= by1

        def begin_drag_select(event) -> str:  # noqa: ANN001 - tkinter event object.
            if render_state["rendering"]:
                return "break"
            start = (preview_canvas.canvasx(event.x), preview_canvas.canvasy(event.y))
            drag_select["start"] = start
            drag_select["moved"] = False
            if drag_select["rect"] is not None:
                preview_canvas.delete(drag_select["rect"])
                drag_select["rect"] = None
            return "break"

        def update_drag_select(event) -> str:  # noqa: ANN001 - tkinter event object.
            start = drag_select.get("start")
            if start is None or render_state["rendering"]:
                return "break"
            start_x, start_y = start
            current_x = preview_canvas.canvasx(event.x)
            current_y = preview_canvas.canvasy(event.y)
            if abs(current_x - start_x) < 5 and abs(current_y - start_y) < 5:
                return "break"
            drag_select["moved"] = True
            rect = drag_select.get("rect")
            if rect is None:
                rect = preview_canvas.create_rectangle(
                    start_x,
                    start_y,
                    current_x,
                    current_y,
                    outline="#0b72d9",
                    dash=(4, 2),
                    width=2,
                    tags=("drag_select_rect",),
                )
                drag_select["rect"] = rect
            else:
                preview_canvas.coords(rect, start_x, start_y, current_x, current_y)
            return "break"

        def finish_drag_select(event) -> str:  # noqa: ANN001 - tkinter event object.
            start = drag_select.get("start")
            if start is None or render_state["rendering"]:
                return "break"
            start_x, start_y = start
            end_x = preview_canvas.canvasx(event.x)
            end_y = preview_canvas.canvasy(event.y)
            if drag_select.get("moved"):
                selection = (
                    min(start_x, end_x),
                    min(start_y, end_y),
                    max(start_x, end_x),
                    max(start_y, end_y),
                )
                changed_selection = False
                for page_index, bounds in page_bounds.items():
                    if rectangles_intersect(selection, bounds):
                        selected_pages.add(page_index)
                        page_vars[page_index].set(True)
                        changed_selection = True
                if changed_selection:
                    set_range_text_from_selection()
            else:
                page_index = page_at(end_x, end_y)
                if page_index is not None:
                    toggle_from_page(page_index, page_vars[page_index])
            if drag_select["rect"] is not None:
                preview_canvas.delete(drag_select["rect"])
            drag_select["start"] = None
            drag_select["rect"] = None
            drag_select["moved"] = False
            return "break"

        preview_canvas.bind("<ButtonPress-1>", begin_drag_select)
        preview_canvas.bind("<B1-Motion>", update_drag_select)
        preview_canvas.bind("<ButtonRelease-1>", finish_drag_select)

        def cancel_render_job() -> None:
            if render_state["after"] is not None:
                try:
                    root.after_cancel(render_state["after"])
                except Exception:
                    pass
                render_state["after"] = None

        def destroy_checkbox_widgets() -> None:
            for checkbox in checkbox_widgets:
                try:
                    checkbox.destroy()
                except Exception:
                    pass
            checkbox_widgets.clear()

        def render_grid() -> None:
            if not page_vars:
                return
            cancel_render_job()
            available_width = max(260, preview_canvas.winfo_width() - 28)
            gap_x = 22
            gap_y = 24
            min_tile_width = 130
            default_tile_width = 260
            max_columns = max(1, min(len(doc), int((available_width + gap_x) // (min_tile_width + gap_x))))
            auto_columns = max(1, min(max_columns, int((available_width + gap_x) // (default_tile_width + gap_x))))
            requested_columns = thumbnail_zoom["columns"]
            columns = requested_columns if requested_columns is not None else auto_columns
            columns = max(1, min(max_columns, len(doc), columns))
            tile_width = int((available_width - gap_x * (columns - 1)) / columns)
            tile_width = max(1, tile_width)
            max_preview_width = tile_width
            label_height = 34

            if (
                layout_state["columns"] == columns
                and abs(layout_state["tile_width"] - tile_width) < 4
                and preview_canvas.find_withtag("page_item")
                and not render_state["rendering"]
            ):
                return

            page_heights: list[int] = []
            for page_index in range(len(doc)):
                rect = doc.load_page(page_index).rect
                page_height = int(max(1, rect.height) * max_preview_width / max(1, rect.width))
                page_heights.append(max(1, page_height))

            rows = (len(doc) + columns - 1) // columns
            row_tops: list[int] = []
            current_y = 14
            for row in range(rows):
                start = row * columns
                end = min(len(doc), start + columns)
                tallest_page = max(page_heights[start:end], default=1)
                row_tops.append(current_y)
                current_y += tallest_page + label_height + gap_y

            layout_state["columns"] = columns
            layout_state["tile_width"] = tile_width
            render_state["token"] += 1
            token = render_state["token"]
            render_state["rendering"] = True

            destroy_checkbox_widgets()
            preview_canvas.delete("page_item")
            preview_canvas.delete("loading_item")
            preview_canvas.delete("drag_select_rect")
            thumbnails.clear()
            page_bounds.clear()
            set_preview_loading(True)

            total_width = 28 + columns * tile_width + (columns - 1) * gap_x
            total_height = current_y - gap_y + 14 if rows else 0
            preview_canvas.configure(scrollregion=(0, 0, total_width, total_height))
            preview_canvas.create_text(
                max(preview_canvas.winfo_width() // 2, 160),
                max(preview_canvas.winfo_height() // 2, 80),
                text=tr("split_loading"),
                anchor="center",
                fill="#555555",
                font=("Microsoft YaHei UI", 14),
                tags=("loading_item",),
            )

            def render_page_batch(start_index: int) -> None:
                if token != render_state["token"] or not root.winfo_exists():
                    return
                preview_canvas.delete("loading_item")
                batch_size = 8
                end_index = min(len(doc), start_index + batch_size)

                try:
                    for index in range(start_index, end_index):
                        page = doc.load_page(index)
                        rect = page.rect
                        zoom = max_preview_width / max(1, rect.width)
                        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
                        image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
                        photo = ImageTk.PhotoImage(image)
                        thumbnails.append(photo)
                        var = page_vars[index]

                        row = index // columns
                        column = index % columns
                        x = 14 + column * (tile_width + gap_x)
                        y = row_tops[row]
                        tag = f"page_{index}"
                        page_bounds[index] = (x, y, x + max(1, pix.width) - 1, y + max(1, pix.height) - 1)

                        preview_canvas.create_image(
                            x,
                            y,
                            image=photo,
                            anchor="nw",
                            tags=("page_item", tag),
                        )
                        preview_canvas.create_rectangle(
                            x,
                            y,
                            x + max(1, pix.width) - 1,
                            y + max(1, pix.height) - 1,
                            fill="",
                            outline="#222222",
                            width=1,
                            tags=("page_item", tag),
                        )
                        checkbox = ttk.Checkbutton(
                            preview_canvas,
                            variable=var,
                            command=lambda page_index=index, variable=var: sync_selection(
                                page_index,
                                variable,
                                True,
                            ),
                        )
                        checkbox_widgets.append(checkbox)
                        preview_canvas.create_window(
                            x + 6,
                            y + 6,
                            window=checkbox,
                            anchor="nw",
                            tags=("page_item",),
                        )
                        preview_canvas.create_text(
                            x,
                            y + pix.height + 10,
                            text=tr("page_label", page=index + 1),
                            anchor="nw",
                            fill="#111111",
                            font=("Microsoft YaHei UI", 11),
                            tags=("page_item", tag),
                        )
                except Exception:
                    render_state["after"] = None
                    render_state["rendering"] = False
                    set_preview_loading(False)
                    messagebox.showwarning(tr("split_pdf"), tr("split_preview_failed"))
                    return

                if end_index < len(doc):
                    render_state["after"] = root.after(1, lambda: render_page_batch(end_index))
                    return

                render_state["after"] = None
                render_state["rendering"] = False
                set_preview_loading(False)
                if total_height <= preview_canvas.winfo_height():
                    preview_canvas.yview_moveto(0)

            render_state["after"] = root.after(10, lambda: render_page_batch(0))

        try:
            root.update_idletasks()
            render_grid()
        except Exception:
            cancel_render_job()
            if resize_job["id"] is not None:
                try:
                    root.after_cancel(resize_job["id"])
                except Exception:
                    pass
                resize_job["id"] = None
            destroy_checkbox_widgets()
            split_frame.destroy()
            for widget in managed_widgets:
                widget.grid()
            root.title(previous_title)
            doc.close()
            messagebox.showwarning(tr("split_pdf"), tr("split_preview_failed"))
            return None

        root.wait_variable(done_var)
        render_state["token"] += 1
        cancel_render_job()
        if resize_job["id"] is not None:
            try:
                root.after_cancel(resize_job["id"])
            except Exception:
                pass
            resize_job["id"] = None
        unbind_mousewheel()
        destroy_checkbox_widgets()
        split_frame.destroy()
        for widget in managed_widgets:
            widget.grid()
        root.title(previous_title)
        doc.close()
        return result

    def split_selected_pdf() -> None:
        indices = selected_indices()
        source: Path | None = None
        for idx in indices:
            candidate = selected[idx]
            if candidate.suffix.lower() == ".pdf":
                source = candidate
                break
        if source is None:
            messagebox.showinfo(tr("split_pdf"), tr("split_no_pdf"))
            chosen = filedialog.askopenfilename(
                title=tr("split_choose_pdf"),
                filetypes=[("PDF", "*.pdf"), (tr("all_files"), "*.*")],
            )
            if not chosen:
                return
            source = Path(chosen)

        try:
            output_dir = Path(output_var.get()).expanduser().resolve()
            target = safe_pdf_name(Path(f"{source.stem}_split.pdf"), output_dir)
            page_indices = choose_pdf_pages_with_preview(source)
            if page_indices is None:
                return
            page_count = split_pdf_indices(source, target, page_indices)
            append_log(tr("split_done", count=page_count, path=target))
            status_var.set(tr("split_done", count=page_count, path=target))
            open_conversion_outputs([target], output_dir)
        except Exception as exc:  # noqa: BLE001 - GUI should show friendly errors.
            messagebox.showerror(tr("split_pdf"), tr("split_failed", error=exc))

    drag_state = {"index": None}

    def begin_list_drag(event) -> None:  # noqa: ANN001 - tkinter event object.
        if selected:
            drag_state["index"] = file_list.nearest(event.y)

    def drag_list_item(event) -> None:  # noqa: ANN001 - tkinter event object.
        old_index = drag_state.get("index")
        if old_index is None or not selected:
            return
        new_index = file_list.nearest(event.y)
        if new_index == old_index or not (0 <= new_index < len(selected)):
            return
        item = selected.pop(old_index)
        selected.insert(new_index, item)
        drag_state["index"] = new_index
        refresh_list()
        file_list.selection_clear(0, tk.END)
        file_list.selection_set(new_index)
        file_list.activate(new_index)

    def end_list_drag(_event=None) -> None:  # noqa: ANN001 - tkinter callback.
        drag_state["index"] = None

    file_list.bind("<ButtonPress-1>", begin_list_drag, add="+")
    file_list.bind("<B1-Motion>", drag_list_item, add="+")
    file_list.bind("<ButtonRelease-1>", end_list_drag, add="+")

    move_up_button = ttk.Button(side, command=lambda: move_selection(-1))
    move_up_button.grid(row=0, column=0, sticky="ew", pady=(0, 8))
    move_down_button = ttk.Button(side, command=lambda: move_selection(1))
    move_down_button.grid(row=1, column=0, sticky="ew", pady=(0, 8))
    remove_button = ttk.Button(side, command=remove_selection)
    remove_button.grid(row=2, column=0, sticky="ew")
    split_button = ttk.Button(side, command=split_selected_pdf)
    split_button.grid(row=3, column=0, sticky="ew", pady=(8, 0))

    if DND_FILES:
        def on_drop(event) -> None:  # noqa: ANN001 - tkinter event object.
            paths = [Path(p) for p in root.tk.splitlist(event.data)]
            add_paths(paths)

        file_list.drop_target_register(DND_FILES)
        file_list.dnd_bind("<<Drop>>", on_drop)

    log = tk.Text(main, height=5, wrap="word")
    log.grid(row=1, column=0, columnspan=3, sticky="nsew", pady=(10, 0))
    log.configure(state="disabled")

    bottom = ttk.Frame(root, padding=(14, 0, 14, 14))
    bottom.grid(row=4, column=0, sticky="ew")
    bottom.columnconfigure(0, weight=1)
    ttk.Label(bottom, textvariable=status_var).grid(row=0, column=0, sticky="w")

    start_button = ttk.Button(bottom)
    start_button.grid(row=0, column=1, sticky="e", padx=(8, 0))
    start_merged_button = ttk.Button(bottom)
    start_merged_button.grid(row=0, column=2, sticky="e", padx=(8, 0))

    def apply_language() -> None:
        lang = current_language()
        current_profile_key = profile_key_from_display(type_var.get())
        root.title(tr("title"))
        add_files_button.configure(text=tr("add_files"))
        add_folder_button.configure(text=tr("add_folder"))
        clear_button.configure(text=tr("clear"))
        recursive_check.configure(text=tr("include_subfolders"))
        language_label.configure(text=tr("language"))
        convert_type_label.configure(text=tr("convert_type"))
        merged_name_label.configure(text=tr("merged_name"))
        output_label.configure(text=tr("output_folder"))
        browse_button.configure(text=tr("browse"))
        move_up_button.configure(text=tr("move_up"))
        move_down_button.configure(text=tr("move_down"))
        remove_button.configure(text=tr("remove"))
        split_button.configure(text=tr("split_pdf"))
        start_button.configure(text=tr("start"))
        start_merged_button.configure(text=tr("start_merged"))
        type_menu.configure(values=[PROFILE_LABELS[lang][key] for key in TYPE_PROFILES])
        type_var.set(PROFILE_LABELS[lang][current_profile_key])
        refresh_list()

    def on_language_changed(_event=None) -> None:  # noqa: ANN001 - tkinter callback.
        selected_display = language_display_var.get()
        for code, display in LANGUAGES.items():
            if display == selected_display:
                language_code_var.set(code)
                save_settings({"language": code})
                apply_language()
                return

    language_menu.bind("<<ComboboxSelected>>", on_language_changed)
    apply_language()

    def append_log(line: str) -> None:
        if closing_var.get() or not root.winfo_exists():
            return
        log.configure(state="normal")
        log.insert(tk.END, line + "\n")
        log.see(tk.END)
        log.configure(state="disabled")

    def safe_after(delay_ms: int, callback, *args) -> None:
        if not closing_var.get() and root.winfo_exists():
            root.after(delay_ms, callback, *args)

    def has_active_worker() -> bool:
        active_workers[:] = [thread for thread in active_workers if thread.is_alive()]
        return bool(active_workers)

    def set_start_buttons_state(state: str) -> None:
        start_button.configure(state=state)
        start_merged_button.configure(state=state)

    def finish_close_when_ready() -> None:
        if has_active_worker() and time.monotonic() < close_deadline["time"]:
            root.after(200, finish_close_when_ready)
            return
        try:
            root.quit()
            root.destroy()
        except Exception:
            pass

    def on_close() -> None:
        if closing_var.get():
            return
        closing_var.set(True)
        SHUTDOWN_EVENT.set()
        terminate_tracked_processes()
        close_deadline["time"] = time.monotonic() + 4
        set_start_buttons_state("disabled")
        if has_active_worker():
            status_var.set("正在结束当前转换任务，完成清理后关闭...")
            root.after(200, finish_close_when_ready)
        else:
            finish_close_when_ready()

    def worker(merge_after_convert: bool) -> None:
        def report(result: ConvertResult) -> None:
            icon = tr("ok") if result.ok else tr("fail")
            output = f" -> {result.output}" if result.output else ""
            safe_after(0, append_log, f"[{icon}] {result.source}{output}: {result.message}")

        try:
            inputs = list(selected)
            out = Path(output_var.get()).expanduser().resolve()
            if not inputs:
                safe_after(0, messagebox.showwarning, tr("no_files_title"), tr("no_files_body"))
                return
            merged_name = merged_name_var.get().strip() or "converted.pdf"
            if not merged_name.lower().endswith(".pdf"):
                merged_name += ".pdf"
            safe_after(0, append_log, tr("starting"))
            results = batch_convert(
                inputs,
                out,
                recursive=recursive_var.get(),
                allowed_exts=TYPE_PROFILES.get(profile_key_from_display(type_var.get()), SUPPORTED_EXTS),
                merge_all=merge_after_convert,
                merged_name=merged_name,
                on_result=report,
            )
            ok_count = sum(1 for r in results if r.ok)
            fail_count = len(results) - ok_count
            output_paths = [r.output for r in results if r.ok and r.output]
            safe_after(0, status_var.set, tr("done_status", ok=ok_count, fail=fail_count))
            if fail_count == 0 and output_paths:
                safe_after(0, open_conversion_outputs, output_paths, out)
            safe_after(
                0,
                messagebox.showinfo,
                tr("finished_title"),
                tr("finished_body", ok=ok_count, fail=fail_count),
            )
        finally:
            safe_after(0, set_start_buttons_state, "normal")
            if closing_var.get():
                try:
                    root.after(0, finish_close_when_ready)
                except Exception:
                    pass

    def start(merge_after_convert: bool = False) -> None:
        if closing_var.get():
            return
        SHUTDOWN_EVENT.clear()
        set_start_buttons_state("disabled")
        thread = threading.Thread(target=worker, args=(merge_after_convert,), daemon=True)
        active_workers.append(thread)
        thread.start()

    start_button.configure(command=lambda: start(False))
    start_merged_button.configure(command=lambda: start(True))
    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Batch convert files to PDF.")
    parser.add_argument("inputs", nargs="*", help="Files or folders. Omit to open the app window.")
    parser.add_argument(
        "-o",
        "--out",
        default=str(default_output_dir()),
        help="Output folder. Default: Downloads.",
    )
    parser.add_argument("--no-recursive", action="store_true", help="Do not include subfolders.")
    parser.add_argument("--type", choices=list(TYPE_PROFILES.keys()), default="All supported files")
    parser.add_argument("--merge-images", action="store_true", help="Merge images into one PDF.")
    parser.add_argument("--merge-all", action="store_true", help="Merge all converted PDFs into one PDF.")
    parser.add_argument("--merged-name", default="merged_images.pdf", help="Merged PDF file name.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    SHUTDOWN_EVENT.clear()
    args = parse_args(argv or sys.argv[1:])
    if not args.inputs:
        run_gui()
        return 0

    results = batch_convert(
        [Path(p) for p in args.inputs],
        Path(args.out),
        recursive=not args.no_recursive,
        allowed_exts=TYPE_PROFILES[args.type],
        merge_images=args.merge_images,
        merge_all=args.merge_all,
        merged_name=args.merged_name,
        on_result=lambda r: print(
            f"[{'OK' if r.ok else 'FAIL'}] {r.source}"
            f"{' -> ' + str(r.output) if r.output else ''}: {r.message}"
        ),
    )
    return 0 if all(r.ok for r in results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
